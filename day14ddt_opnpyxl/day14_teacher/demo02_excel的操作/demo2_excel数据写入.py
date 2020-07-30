"""
============================
Author:柠檬班-木森
Time:2020/7/30   21:15
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""
"""
第一步：将excel文件加载到一个工作簿对象中
    wb = openpyxl.load_workbook("文件名")
第二步：选择文件中的表单
    sh = wb[表单名]
第三步：往表格中写入数据
    sh.cell(row=行,column=列,value='写入的值')
# 第四步：保存
    wb.save("文件名")

注意：在往excel中写入数据的时候，文件不能处于一个打开的状态
"""

import openpyxl

# 第一步：将excel文件加载到一个工作簿对象中
wb = openpyxl.load_workbook("musen.xlsx")

# 第二步：选择文件中的表单
sh = wb["login"]

# 第三步：往表格中写入数据
sh.cell(row=1, column=2, value='木森6666')

# 第四步：保存
wb.save("musen.xlsx")
