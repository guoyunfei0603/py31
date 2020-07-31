"""
============================
Author:小白31
Time:2020/8/1 0:33
E-mail:1359239107@qq.com
============================
"""
'''
excel操作三大对象：
工作簿：一个excel文件会被加载为一个工作簿对象Workbook
工作表：excel文件每个sheet表达会被加载成一个工作表对象(sheet)
单元格：工作表中的每一个格子就是一个表格对象（cell）

pip install 包名 -i https://pypi.tuna.tsinghua.edu.cn/simple/

# 注意：在桌面创建.xlsx文件 不要在excel里面直接创建！
'''
import openpyxl
# 1. 创建一个工作簿对象
wb = openpyxl.load_workbook("testdata.xlsx")

# 2. 选择sheet
sheet = wb["register"]

# 3. 选择单元格cell ,行==列
cell = sheet.cell(3,4).value
print(cell)

#----------批量读取---------------
res = list(sheet.rows)
# li = []
for i in res:
    for j in i:
        # li.append(j.value)
        print(j.value,end='   ')
    print()
# print(li)

# ---------写入数据---------------
sheet2 = wb["Sheet2"]
sheet2.cell(row =1, column=1, value="小白呀")

wb.save("testdata.xlsx")

