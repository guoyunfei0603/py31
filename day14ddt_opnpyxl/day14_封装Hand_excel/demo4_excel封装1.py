"""
============================
Author:小白31
Time:2020/8/2 21:42
E-mail:1359239107@qq.com
============================
"""
import openpyxl


class HandExcel:
    def __init__(self,filename,sheetname):
        self.filename = filename
        self.sheetname = sheetname

    # def open_excel(self,filename,sheetname):
    #     wb = openpyxl.load_workbook(filename)
    #
    #     sh = wb[sheetname]

    def read_excel(self):
        """读取excel表格数据"""
        wb = openpyxl.load_workbook(self.filename)

        sh = wb[self.sheetname]

        res = list(sh.rows)

        # 获取第一行的title
        title = []
        for i in res[0]:
            title.append(i.value)

        case_data = []  # 存放用例数据
        # 获取除第一行以外的所有行数据
        for i in res[1:]:
            data = []
            for j in i:
                data.append(j.value)
            res = dict(zip(title, data))
            case_data.append(res)
        return case_data

    def write_excel(self,i,j,value):
        wb = openpyxl.load_workbook(self.filename)

        sh = wb[self.sheetname]
        sh.cell(row=i, column=j, value=value)

        wb.save(self.filename)

excel = HandExcel("test_data.xlsx", "login")

res = excel.read_excel()
print(res)

excel.write_excel(7,6,"写入小白")