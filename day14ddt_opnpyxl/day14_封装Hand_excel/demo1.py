"""
============================
Author:小白31
Time:2020/8/2 21:20
E-mail:1359239107@qq.com
============================
"""
import openpyxl

wb = openpyxl.load_workbook("test_data.xlsx")

sh = wb["login"]
# 读取单元格的值
res = sh.cell(4,2).value
print(res)
# 往指定单元格写入值
sh.cell(row=6, column=1, value="测试")

# 保存
wb.save("case_data.xlsx")