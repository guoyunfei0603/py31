"""
============================
Author:柠檬班-木森
Time:2020/7/30   21:50
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
============================
"""
import openpyxl


class Excel:
    def __init__(self, filename, sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name

    def read_data(self):
        """读数据"""
        # 第一步：将excel文件加载到一个工作簿对象中
        wb = openpyxl.load_workbook(self.filename)
        # 第二步：选择文件中的表单
        sh = wb[self.sheet_name]
        res = list(sh.rows)
        # 获取第一行的单元格
        title = []
        # 遍历第一行所有的单元格，将格子中的值添加到列表title中
        for c in res[0]:
            title.append(c.value)
        cases_data = []
        # 遍历除第一行之外所有的行
        for row in res[1:]:
            data = []
            for c in row:
                data.append(c.value)
            case = dict(zip(title, data))
            # print(case)
            cases_data.append(case)
        return cases_data

    def write_data(self, row, column, value):
        """写数据"""
        # 第一步：将excel文件加载到一个工作簿对象中
        wb = openpyxl.load_workbook(self.filename)
        # 第二步：选择文件中的表单
        sh = wb[self.sheet_name]
        sh.cell(row=row, column=column, value=value)
        wb.save(self.filename)


if __name__ == '__main__':
    excel = Excel("cases.xlsx", "register")
    data = excel.read_data()
    print(data)
